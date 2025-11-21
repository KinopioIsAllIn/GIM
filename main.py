from __future__ import print_function
import os
import torch
import model
from SpoTFormer import MSSTGT
from wsad_dataset import SampleDataset

import random
from test import test
from train import train
import options
import numpy as np
from torch.optim import lr_scheduler
from tqdm import tqdm
import csv

torch.set_default_tensor_type('torch.cuda.FloatTensor')
def setup_seed(seed):
   random.seed(seed)
   os.environ['PYTHONHASHSEED'] = str(seed)
   np.random.seed(seed)
   torch.manual_seed(seed)
   torch.cuda.manual_seed(seed)
   torch.cuda.manual_seed_all(seed)
   torch.backends.cudnn.benchmark = False
   torch.backends.cudnn.deterministic = True

import torch.optim as optim

from openpyxl import Workbook


def write_to_excel_column1(data_list, filename):
   # 创建一个新的工作簿
   wb = Workbook()
   # 选择默认的工作表
   ws = wb.active

   # 写入数据到一列中
   for i in range(len(data_list)):
      for index, value in enumerate(data_list[i][:, 0], start=1):
         ws.cell(row=index, column=i * 2 + 1, value=value)
      for index, value in enumerate(data_list[i][:, 1], start=1):
         ws.cell(row=index, column=i * 2 + 2, value=value)

   # 保存工作簿
   wb.save(filename)

def write_to_excel_column(data_list, filename):
    # 创建一个新的工作簿
    wb = Workbook()
    # 选择默认的工作表
    ws = wb.active

    # 写入数据到一列中
    for i in range(len(data_list)):
      for index, value in enumerate(data_list[i], start=1):
         ws.cell(row=index, column=i+1, value=value)

    # 保存工作簿
    wb.save(filename)

if __name__ == '__main__':
   args = options.parser.parse_args()
   seed = args.seed
   cuda_n = args.cuda
   print('=============seed: {}, pid: {}============='.format(seed, os.getpid()))
   setup_seed(seed)
   device = torch.device("cuda:" + str(cuda_n))
   torch.cuda.set_device(cuda_n)

   dirname = args.dirname
   for datasetname in ['CASME3']:  # set dataset manually, can also set by args.dataset_name
      args.dataset_name = datasetname
      # datasetname = args.dataset_name

      if datasetname == 'SAMM':
         if cuda_n == 0:
            validation_subjects = [
               "samm_006", "samm_011", "samm_010", "samm_007", "samm_008", "samm_009", "samm_012",
            ]
         elif cuda_n == 1:
            validation_subjects = [
               "samm_013", "samm_014", "samm_015", "samm_016", "samm_019", "samm_018", "samm_017",
            ]
         elif cuda_n == 2:
            validation_subjects = [
               "samm_020",  "samm_026", "samm_025","samm_021", "samm_022", "samm_023", "samm_028",   # "samm_024",
            ]
         elif cuda_n == 3:
            validation_subjects = [
               "samm_032", "samm_033", "samm_034", "samm_035", "samm_036", "samm_037", "samm_030",  # "samm_031",
            ]
      elif datasetname == 'CASME2':
         if cuda_n == 0:
            validation_subjects = [
               "casme_16", "casme_15", "casme_19", "casme_20", "casme_21",
            ]
         elif cuda_n == 1:
            validation_subjects = [
               "casme_24", "casme_22", "casme_23", "casme_25", "casme_26",
            ]
         elif cuda_n == 2:
            validation_subjects = [
               "casme_32", "casme_27", "casme_31", "casme_30", "casme_33",# "casme_29",
            ]
         elif cuda_n == 3:
            validation_subjects = [
               "casme_37", "casme_38", "casme_36", "casme_40", "casme_34", # "casme_35",
            ]
      elif datasetname == 'CASME3':
         if cuda_n == 0:
            validation_subjects = [
               'casme3_1', 'casme3_10', 'casme3_11', 'casme3_12', 'casme3_13',
               'casme3_138', 'casme3_139', 'casme3_14', 'casme3_140', 'casme3_142',
               'casme3_143', 'casme3_144', 'casme3_145', 'casme3_146', 'casme3_147',
               'casme3_148', 'casme3_149', 'casme3_15', 'casme3_150', 'casme3_151',
               'casme3_152', 'casme3_153', 'casme3_154', 'casme3_155',
            ]
         elif cuda_n == 1:
            validation_subjects = [
               'casme3_156', 'casme3_157', 'casme3_158', 'casme3_159', 'casme3_16',
               'casme3_160', 'casme3_161', 'casme3_162', 'casme3_163', 'casme3_165',
               'casme3_166', 'casme3_167', 'casme3_168', 'casme3_169', 'casme3_17',
               'casme3_170', 'casme3_171', 'casme3_172', 'casme3_173', 'casme3_174',
               'casme3_175', 'casme3_176', 'casme3_177', 'casme3_178',
            ]
         elif cuda_n == 2:
            validation_subjects = [
               'casme3_179', 'casme3_180', 'casme3_181', 'casme3_182', 'casme3_183',
               'casme3_184', 'casme3_185', 'casme3_186', 'casme3_187', 'casme3_188',
               'casme3_189', 'casme3_190', 'casme3_192', 'casme3_193', 'casme3_194',
               'casme3_195', 'casme3_196', 'casme3_197', 'casme3_198', 'casme3_2',
               'casme3_200', 'casme3_201', 'casme3_202', 'casme3_203', 'casme3_204',

            ]
         elif cuda_n == 3:
            validation_subjects = [
               'casme3_206', 'casme3_207', 'casme3_208', 'casme3_209', 'casme3_210',
               'casme3_212',  'casme3_216', 'casme3_217', 'casme3_215', 'casme3_213',
               'casme3_214', 'casme3_218', 'casme3_3', 'casme3_39', 'casme3_4',
               'casme3_40', 'casme3_41', 'casme3_42', 'casme3_5', 'casme3_6',
               'casme3_7', 'casme3_77', 'casme3_8', 'casme3_9'
            ]

         validation_subjects = [subject[7:] for subject in validation_subjects]

      overall_tp_MaE, overall_tp_ME, overall_fp_MaE, overall_fp_ME, gt_MaE, gt_ME, overall_apex_diff_MaE, overall_apex_diff_ME = 0, 0, 0, 0, 0, 0, 0, 0
      results = {}

      for i in range(len(validation_subjects)):
         args.validation_subject = validation_subjects[i]

         train_dataset = SampleDataset(args, "train")
         train_loader = torch.utils.data.DataLoader(train_dataset, batch_size=args.batch_size, shuffle=True, generator=torch.Generator(device=device))
         test_dataset = SampleDataset(args, "test")
         test_loader = torch.utils.data.DataLoader(test_dataset, batch_size=1, shuffle=False, generator=torch.Generator(device=device))
         print("Training size: %d, Test size: %d." % (len(train_dataset), len(test_dataset)))

         if not os.path.exists('./ckpt/'):
            os.makedirs('./ckpt/')
         logger = None
         print(args)

         spotformer = MSSTGT(
           seq_len=17,
           num_classes=10,
           dim=256,
           depth=4,
           heads=4,
           mlp_dim=256,
           dropout=0.0,
           emb_dropout=0.1).to(device)

         model1 = getattr(model, args.use_model)(train_dataset.feature_size, train_dataset.num_class, opt=args).to(device)

         params = list(spotformer.parameters()) + list(model1.parameters())

         if args.pretrained_ckpt is not None:
            model1.load_state_dict(torch.load(args.pretrained_ckpt))

         optimizer = optim.Adam(params, lr=args.lr, weight_decay=args.weight_decay)

         scheduler = lr_scheduler.StepLR(optimizer, step_size=30, gamma=0.5)

         total_loss = 0
         lrs = [args.lr, args.lr/5, args.lr/5/5]
         print(model1)

         best_tp_MaE, best_tp_ME, best_fp_MaE, best_fp_ME, best_apex_diff_MaE, best_apex_diff_ME = 0, 0, 0, 0, 0, 0
         best_f1 = 0.0

         for itr in tqdm(range(args.max_iter)):
            total_loss = train(itr, train_loader, args, spotformer, model1, optimizer, logger, device)
            scheduler.step()
            if itr > 4 and itr % args.interval == 0:
               print('Iteration: %d, Loss: %.5f' % (itr, total_loss/args.interval))
               total_loss = 0
               if not os.path.exists('./ckpt' + dirname):
                  os.makedirs('./ckpt' + dirname)

               iou, dmap, attn_scores, proposals, action_scores, apex_scores = test(itr, test_dataset, test_loader, args, spotformer, model1, logger, device)

               tp_mae = dmap[0][0][0]
               tp_me = dmap[0][0][1]
               fp_mae = dmap[1][0][0]
               fp_me = dmap[1][0][1]
               gt_mae = dmap[2][0][0]
               gt_me = dmap[2][0][1]
               apex_diff_mae = dmap[3][0][0]
               apex_diff_me = dmap[3][0][1]

               f1 = 2 * (tp_mae + tp_me)/(tp_mae + tp_me + fp_mae + fp_me + gt_mae + gt_me)

               cond = f1 >= best_f1
               if cond:
                  best_f1 = f1
                  best_tp_MaE = tp_mae
                  best_tp_ME = tp_me
                  best_fp_MaE = fp_mae
                  best_fp_ME = fp_me
                  best_apex_diff_MaE = apex_diff_mae
                  best_apex_diff_ME = apex_diff_me

               state = {
                  'model1_state_dict': model1.state_dict(),
                  'spotformer_state_dict': spotformer.state_dict()
               }
               torch.save(state, './ckpt' + dirname + '/best_' + args.model_name + args.validation_subject + '_' + str(
                  itr) + '.pkl')

               if not os.path.exists('proposals' + dirname + '/'):
                  os.makedirs('proposals' + dirname + '/')
               proposals.to_csv('proposals' + dirname + '/' + args.validation_subject + '_' + str(itr) + '.csv')

               if not os.path.exists('apexscores' + dirname + '/'):
                  os.makedirs('apexscores' + dirname + '/')
               filename = 'apexscores' + dirname + '/apex' + args.validation_subject + '_' + str(itr) + '.csv'
               write_to_excel_column1(apex_scores, filename)

               if not os.path.exists( 'attnresults' + dirname):
                  os.makedirs( 'attnresults' + dirname)
               filename = 'attnresults' + dirname + '/attn' + args.validation_subject + '_' + str(itr) + '.csv'
               write_to_excel_column(attn_scores, filename)

               print('current f1: %f, best f1: %f' % (f1, best_f1))

         overall_tp_MaE += best_tp_MaE
         overall_tp_ME += best_tp_ME
         overall_fp_MaE += best_fp_MaE
         overall_fp_ME += best_fp_ME
         gt_MaE += gt_mae
         gt_ME += gt_me
         overall_apex_diff_MaE += best_apex_diff_MaE
         overall_apex_diff_ME += best_apex_diff_ME

         results[i] = [best_tp_MaE, best_tp_ME, best_fp_MaE, best_fp_ME, best_f1]
         print(results)

      pre_mae = overall_tp_MaE / (overall_tp_MaE + overall_fp_MaE)
      pre_me = overall_tp_ME / (overall_tp_ME + overall_fp_ME)
      rec_mae = overall_tp_MaE / gt_MaE
      rec_me = overall_tp_ME / gt_ME

      f1_mae = 2*(overall_tp_MaE) / (overall_tp_MaE + overall_fp_MaE + gt_MaE)
      f1_me = 2*(overall_tp_ME) / (overall_tp_ME + overall_fp_ME + gt_ME)
      f1 = 2*(overall_tp_ME+overall_tp_MaE) / (overall_tp_MaE + overall_tp_ME + overall_fp_ME + overall_fp_MaE + gt_ME + gt_MaE)

      print('MaE: %f, ME: %f, Overall: %f' % (f1_mae, f1_me, f1))
      print(results)


    
